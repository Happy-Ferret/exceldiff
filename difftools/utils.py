from Netease.settings import BASE_DIR
from openpyxl import Workbook
from openpyxl import load_workbook
import difflib
import json

def handle_uploaded_file(f, num):
    ext = "xlsx"
    filename = f.name
    filext = filename.split(".")
    if len(filext) > 1:
        ext = filext[1]
    with open(BASE_DIR + "/media/" + str(num) + "." + ext, 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
    return ext

def diffpoints(data1, data2, datadiff, row_insert, col_insert, row_delete, col_delete):
    row = 0
    for d1, d2 in zip(data1, data2):
        if not (row in row_insert or row in row_delete):
            col = 0
            for ele1, ele2 in zip(d1, d2):
                if not (col in col_insert or col in col_delete):
                    if ele1 != ele2:
                        datadiff.append([row, col])
                col = col + 1
        row = row + 1


def gendict(dict, data1, data2, row_insert, col_insert, row_delete, col_delete, datadiff,
            col_header1, col_header2, row_header1, row_header2):
    dict["data1"] = data1
    dict["data2"] = data2
    dict["row_insert"] = row_insert
    dict["col_insert"] = col_insert
    dict["row_delete"] = row_delete
    dict["col_delete"] = col_delete
    dict["datadiff"] = datadiff
    dict["col_header1"] = col_header1
    dict["col_header2"] = col_header2
    dict["row_header1"] = row_header1
    dict["row_header2"] = row_header2

#https://stackoverflow.com/questions/23861680/convert-spreadsheet-number-to-column-letter
def column_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def single_column_header(header, length, col):
    cursor = 0
    for i in range(length):
        if i in col:
            cursor = cursor + 1
            header.append("")
        else:
            header.append(column_string(i - cursor + 1))

def single_row_header(header, length, row):
    cursor = 0
    for i in range(length):
        if i in row:
            cursor = cursor + 1
            header.append("")
        else:
            header.append(i - cursor + 1)


def headers(col_header1, col_header2, row_length, col_length, row_header1, row_header2,
            row_insert, col_insert, row_delete, col_delete):
    single_column_header(col_header1, col_length, col_insert)
    single_column_header(col_header2, col_length, col_delete)
    single_row_header(row_header1, row_length, row_insert)
    single_row_header(row_header2, row_length, row_delete)

def gendata(data1, data2, row_length, col_length, row_insert, col_insert, row_delete, col_delete, t1, t2):
    cur = 0
    for i in range(row_length):
        row = []
        data = [''] * col_length
        if i in row_insert or i - cur >= len(t1):
            cur = cur + 1
            data1.append(data)
            continue
        else:

            row = t1[i - cur]
            current = 0
            for j in range(col_length):
                if j in col_insert or j - current >= len(row):
                    current = current + 1
                else:
                    data[j] = row[j - current]
            data1.append(data)
    cur = 0
    for i in range(row_length):
        row = []
        data = [''] * col_length
        if i in row_delete or i - cur >= len(t2):
            cur = cur + 1
            data2.append(data)
            continue
        else:
            row = t2[i - cur]
            current = 0
            for j in range(col_length):
                if j in col_delete or j - current >= len(row):
                    current = current + 1
                else:
                    data[j] = row[j - current]
            data2.append(data)


def opcodes(delete, insert, opcodes):
    cursor_a = 0
    cursor_b = 0
    for tag, i1, i2, j1, j2 in opcodes:
        if tag == 'delete':
            for num in range(i1, i2):
                delete.append(num + cursor_a)
                cursor_b = cursor_b + 1
        elif tag == 'insert':
            for num in range(j1, j2):
                insert.append(num + cursor_b)
                cursor_a = cursor_a + 1
        elif tag == 'replace':
            cursor_b = cursor_b + i2 - i1
            for num in range(i1, i2):
                delete.append(num + cursor_a)
                insert.append(num + cursor_b)
            cursor_a = cursor_a + i2 - i1
    return [cursor_a, cursor_b]

def execute_diff(ext):
    wb1 = load_workbook(filename = BASE_DIR + "/media/" + '1.' + ext)
    wb2 = load_workbook(filename = BASE_DIR + "/media/" + '2.' + ext)

    # grab the active worksheet
    ws1 = wb1.active
    ws2 = wb2.active

    # Data can be assigned directly to cells
    t1 = tuple(ws1.rows)
    t2 = tuple(ws2.rows)

    len1 = len(t1)
    len2 = len(t2)

    arr = []
    for t in t1:
        row = []
        for i in t:
            row.append(i.value)
        tuprow = tuple(row)
        arr.append(tuprow)

    t1 = tuple(arr)

    arr = []
    for t in t2:
        row = []
        for i in t:
            row.append(i.value)
        tuprow = tuple(row)
        arr.append(tuprow)

    t2 = tuple(arr)

    s = difflib.SequenceMatcher(None, t1[0], t2[0])
    longmatch_line = s.find_longest_match(0, len(t1[0]), 0, len(t2[0]))
    line_a = longmatch_line.a
    line_b = longmatch_line.b
    opcode = s.get_opcodes()

    col_delete = []
    col_insert = []
    cursors = opcodes(col_delete, col_insert, opcode)

    col_length = cursors[0] + len(t1[0])

    c1 = tuple(ws1.columns)
    c2 = tuple(ws2.columns)

    arr = []
    for c in c1:
        row = []
        for i in c:
            row.append(i.value)
        tuprow = tuple(row)
        arr.append(tuprow)
    c1 = tuple(arr)

    arr = []
    for c in c2:
        row = []
        for i in c:
            row.append(i.value)
        tuprow = tuple(row)
        arr.append(tuprow)

    c2 = tuple(arr)


    s = difflib.SequenceMatcher(None, c1[line_a], c2[line_b])
    opcode = s.get_opcodes()

    row_delete = []
    row_insert = []
    cursors = opcodes(row_delete, row_insert, opcode)

    row_length = cursors[0] + len(c1[0])

    data1 = []
    data2 = []
    datadiff = []

    col_header1 = []
    col_header2 = []
    row_header1 = []
    row_header2 = []

    gendata(data1, data2, row_length, col_length, row_insert, col_insert, row_delete, col_delete, t1, t2)
    diffpoints(data1, data2, datadiff, row_insert, col_insert, row_delete, col_delete)
    headers(col_header1, col_header2, row_length, col_length, row_header1, row_header2, row_insert, col_insert,
            row_delete, col_delete)
    dict = {}
    gendict(dict, data1, data2, row_insert, col_insert, row_delete, col_delete, datadiff, col_header1, col_header2,
            row_header1, row_header2)
    with open(BASE_DIR + "/media/" + 'result.json', 'w') as fp:
        json.dump(dict, fp)
